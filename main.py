from urllib.request import urlopen

import requests as requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


class Movie:
    def __init__(self, name, date, genres, link):
        self.name = name[0:name.index('(')].strip()
        self.month = date.split(' ')[0]
        ordinal = lambda n: "%d%s" % (
            n, "tsnrhtdd"[(n // 10 % 10 != 1) * (n % 10 < 4) * n % 10::4])  # formats the ordinal suffix of a date
        self.day = ordinal(int(date.split(' ')[1]))
        self.genres = genres
        self.link = link

    def cell_display(self):
        return self.name + " - " + self.genres


class MarvelIP:
    def __init__(self, name, date):
        self.name = name
        self.month = date.split(' ')[0]
        ordinal = lambda n: "%d%s" % (
            n, "tsnrhtdd"[(n // 10 % 10 != 1) * (n % 10 < 4) * n % 10::4])  # formats the ordinal suffix of a date
        self.day = ordinal(int(date.split(' ')[1][0:1:1]))

    def cell_display(self):
        str = self.name + " - " + self.month
        if self.month != 'TBD' and self.month != "Summer":
            str += " " + self.day
        return str


def has_bad_genre(genres):
    for genre in genres:
        if genre.text.strip() in bad_genres:
            return True
    return len(genres) == 1 and genres[0].text.strip() == "Drama"


def set_style(sheet):
    for col in sheet.iter_cols(min_col=1, max_col=16, min_row=1, max_row=100):
        for cell in col:
            cell.font = Font(name='Arial', size=10)
    for x in range(1, 20):
        sheet.column_dimensions[get_column_letter(x)].width = 15


"""
IMDb MOVIE LIST
"""
workbook = Workbook()
sheet = workbook.active
set_style(sheet)
bad_genres = {"Romance", "Family", "Documentary", "Musical", "Biography",
              "History"}  # and movies w/ one genre that's Drama

file_name = 'Movie Chart ' + datetime.now().strftime('%Y.%m.%d - %H.%M.%S') + '.xlsx'
print('Beginning scrub of www.imdb.com/movies-coming-soon/.')
print('Filtered genres: ' + str(bad_genres) + " + 'Drama' by itself")

movie_list = []  # array of type Movie

for x in range(1, 13):
    quote_page = 'https://www.imdb.com/movies-coming-soon/2022-'
    if x < 10:
        quote_page += '0'
    quote_page += str(x)

    page = urlopen(quote_page)  # open the URL
    soup = BeautifulSoup(page, 'html.parser')  # create the BeautifulSoup object
    divs = soup.find(class_='list detail').findAll(class_=['list_item odd', 'list_item even', 'li_group'])

    y = 0
    date_of_movie = None
    while y < len(divs):
        if divs[y].name == 'h4':  # if date
            date_of_movie = divs[y].text.strip()
        elif not has_bad_genre(divs[y].find('p').findAll('span')):  # if movie with no bad genres
            genres = ''
            for genre in divs[y].find('p').findAll('span'):  # adds genres of current movie to string
                genres += genre.text.strip() + " "
            movie_list.append(Movie(divs[y].find('h4').text,
                                    date_of_movie,
                                    genres,
                                    'https://www.imdb.com' + divs[y].find('h4').find('a').get('href')))
        y += 1

max_column_num = 1  # starts at 1 since all movies are shifted over 1 column
max_chars_per_cell = 19  # determines how many chars to count before adding another column to a movie entry
for movie in movie_list:
    if len(movie.cell_display()) >= max_chars_per_cell:
        col_num = 2 + int(len(movie.cell_display()) / max_chars_per_cell)
        if col_num > max_column_num:
            max_column_num = col_num

row = 1
current_month = None  # used to see if we need to add another month entry
current_day = None  # used to compare with previous movie entries to see if date is required
use_color = False
for movie in movie_list:
    loc = 'A' + str(row)
    if current_month != movie.month:
        current_month = movie.month
        sheet[loc] = movie.month
        sheet[loc].font += Font(bold=True)
        sheet[loc].fill = PatternFill(fill_type='solid', start_color='f3cb4b')
        sheet[loc].border += Border(right=Side(style='thin'), top=Side(style='thin'))
        sheet['B' + str(row)].fill = PatternFill(fill_type='solid', start_color='f3cb4b')
        sheet['B' + str(row)].border += Border(right=Side(style='thin'), top=Side(style='thin'))
        sheet.merge_cells('B' + str(row) + ':' + get_column_letter(max_column_num) + str(row))
        row += 1
    loc = 'A' + str(row)
    if current_day != movie.day:
        current_day = movie.day
        sheet[loc] = movie.day
        sheet[loc].font += Font(italic=True)
    if use_color:
        sheet[loc].fill = PatternFill(fill_type='solid', start_color='fdf7e0')
    loc = 'B' + str(row)
    sheet[loc] = movie.cell_display()
    sheet[loc].hyperlink = movie.link
    sheet[loc].border += Border(left=Side(style='thin'), right=Side(style='thin'))
    if use_color:
        sheet[loc].fill = PatternFill(fill_type='solid', start_color='fdf7e0')
    sheet.merge_cells(loc + ':' + get_column_letter(max_column_num) + str(row))
    use_color = ~use_color
    row += 1
sheet['A' + str(row - 1)].border += Border(bottom=Side(style='thin'))
sheet['B' + str(row - 1)].border += Border(bottom=Side(style='thin'))
print('Done!')

"""
MARVEL RELEASE SCHEDULE
"""
print('Beginning scrub of https://www.mcuschedule.com/.')
page = requests.get("https://www.mcuschedule.com/", verify=True)  # open the URL
soup = BeautifulSoup(page.content, 'html.parser')  # create the BeautifulSoup object
divs = soup.findAll(class_='movie-label')

marvel_list = []

for div in divs:
    if div.find('h3').text.__contains__('2023'):
        break
    marvel_list.append(MarvelIP(div.find('h2').text, div.find('h3').text))

# find the maximum col length of the list
max_column_num_2 = 1
for marvelIP in marvel_list:
    if len(marvelIP.cell_display()) >= max_chars_per_cell:
        col_num = 1 + int(len(marvelIP.cell_display()) / max_chars_per_cell)
        if col_num > max_column_num_2:
            max_column_num_2 = col_num

loc = get_column_letter(max_column_num + 1) + '1'
sheet[loc] = "Marvel Releases"
sheet[loc].alignment = Alignment(horizontal='center', vertical='center')
sheet[loc].fill = PatternFill(fill_type='solid', start_color='e7175c')
sheet[loc].border += Border(right=Side(style='thin'))
sheet.merge_cells(
    loc + ':' + get_column_letter(max_column_num + max_column_num_2) + '1')  # make the entire list span that # of cols

row = 2
use_color = False
for marvelIP in marvel_list:
    loc = get_column_letter(max_column_num + 1) + str(row)
    sheet[loc] = marvelIP.cell_display()
    sheet[loc].border += Border(right=Side(style='thin'))
    if use_color:
        sheet[loc].fill = PatternFill(fill_type='solid', start_color='fdd8e5')
    sheet.merge_cells(loc + ':' + get_column_letter(max_column_num + max_column_num_2) + str(row))
    use_color = ~use_color
    row += 1
sheet[get_column_letter(max_column_num + 1) + str(row - 1)].border += Border(bottom=Side(style='thin'))
print('Done!')

print('Created \'' + file_name + '\' in Python script directory!')
workbook.save(filename=file_name)
