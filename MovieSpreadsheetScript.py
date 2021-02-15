from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def int_to_month(num):
    switcher = {
        1: 'January',
        2: 'February',
        3: 'March',
        4: 'April',
        5: 'May',
        6: 'June',
        7: 'July',
        8: 'August',
        9: 'September',
        10: 'October',
        11: 'November',
        12: 'December',
    }
    return switcher.get(num, "error: invalid month number!")


def has_bad_genre(genres):
    if len(genres) == 1 and genres[0].text.strip() == "Drama":
        return True
    for genre in genres:
        if genre.text.strip() in bad_genres:
            return True
    return False


def format_date(date):
    switcher = {
        1: "st",
        2: "nd",
        3: "rd",
        21: "st",
        22: "nd",
        23: "rd",
        31: "st",
    }
    return date + switcher.get(int(date), "th")


def set_style(sheet):
    for col in sheet.iter_cols(min_col=1, max_col=8, min_row=1, max_row=100):
        for cell in col:
            cell.font = Font(name='Arial', size=10)
    x = 1
    while x < 9:
        sheet.column_dimensions[get_column_letter(x)].width = 15
        x += 1

def set_cell_border(min_col, max_col, min_row, max_row, side, is_bottom_month_border):
    merged_cell_found = False
    switcher = {
        'left': Border(left=Side(style='thin')),
        'right': Border(right=Side(style='thin')),
        'top': Border(top=Side(style='thin')),
        'bottom': Border(bottom=Side(style='thin')),
    }
    if min_row == 0 and max_row == 0:
        return

    for col in sheet.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row):
        for cell in col:
            if is_bottom_month_border:
                cell.border += switcher.get(side, 'error: not a valid side')
            else:
                if isinstance(cell, MergedCell):
                    merged_cell_found = True
                    continue
                if merged_cell_found and not isinstance(cell, MergedCell):
                    cell.border += switcher.get(side, 'error: not a valid side')


global row
row = 1

workbook = Workbook()
sheet = workbook.active
set_style(sheet)
bad_genres = {"Romance", "Family", "Documentary", "Musical", "Biography", "History"}  # and movies w/ one genre that's Drama

file_name = 'Movie Chart ' + datetime.now().strftime('%Y.%m.%d - %H.%M.%S') + '.xlsx'
print('Beginning scrubbing process of www.imdb.com/movies-coming-soon/.')
print('Filtered genres: ' + str(bad_genres))

for x in range(1, 13):
    print('Searching for movies in ' + int_to_month(x) + '...')
    cell_loc = 'A' + str(row)
    sheet[cell_loc] = int_to_month(x)
    sheet[cell_loc].font += Font(bold=True)
    if x > 9:
        quote_page = 'https://www.imdb.com/movies-coming-soon/2021-' + str(x) + "/"
    else:
        quote_page = 'https://www.imdb.com/movies-coming-soon/2021-0' + str(x) + "/"
    sheet[cell_loc].hyperlink = quote_page

    page = urlopen(quote_page)
    soup = BeautifulSoup(page, 'html.parser')
    divs = soup.find(class_='list detail').findAll(class_=['list_item odd', 'list_item even', 'li_group'])

    """
    Formats the set 'divs' to only contain dates containing valid movies.
    Divides up 'divs' into subsets from one date to the next. Goes through
    each set and asks if the movie is valid. If all movies in a set are removed,
    the date is consequently removed as well. This shit took too long to figure
    out and it's sloppy as hell. 
    """
    y = 0
    while y < len(divs):
        if divs[y].name == 'h4': # if it's a date
            z = y + 1
            try:
                end_point = divs.index(next(x for x in divs[y + 1:len(divs)] if x.name == 'h4')) # find the next instance of a date
            except StopIteration as e:
                if z == len(divs): # if final entry in divs is a date, remove it
                    divs.remove(divs[y])
                break
            if z == end_point: # if date has no movies associated with it
                divs.remove(divs[y])
                continue
            start_size = len(divs)
            subset_size = len(divs[y + 1:end_point])
            while z < end_point: # iterate through set until next date
                if divs[z].find('h4') is not None and has_bad_genre(divs[z].find('p').findAll('span')): # if not valid movie
                    divs.remove(divs[z])
                    z -= 1
                    end_point -= 1
                if y == z and start_size - len(divs) == subset_size: # if removed every movie from subset
                    divs.remove(divs[z])
                    end_point -= 1
                z += 1
        y = end_point

    current_month_row = row
    max_merged_cell_len = 0
    row += 1

    for y in range(0, len(divs)):
        if divs[y].name == 'h4':
            cell_loc = 'A' + str(row)
            date = divs[y].text.strip()
            print(date)
            sheet[cell_loc] = format_date(date[date.index(' ') + 1:len(date)])
            sheet[cell_loc].font += Font(italic=True)
            continue

        cell_loc = 'B' + str(row)
        movie_name = divs[y].find('h4').text.strip() + ' - '
        print('\tFound movie \'' + movie_name[0:movie_name.index('(') - 1] + '\'!')
        for genre in divs[y].find('p').findAll('span'): # adds genres of current movie to string
            movie_name += genre.text.strip() + " "

        if '(' in movie_name: # checks if movie has a year added to it
            sheet[cell_loc] = movie_name[0:movie_name.index('(') - 1] + movie_name[
                                                                              movie_name.index(')') + 1:len(
                                                                                  movie_name)]
        else:
            sheet[cell_loc] = movie_name
        sheet[cell_loc].hyperlink = 'https://www.imdb.com' + divs[y].find('h4').find('a').get('href')

        # calculates how many columns to merge
        x = len(sheet['B' + str(row)].value)
        num_of_cols = 0
        while x > 0:
            x -= 18
            num_of_cols += 1
        data=sheet[cell_loc].value
        sheet.merge_cells(cell_loc + ':' + get_column_letter(1 + num_of_cols) + str(row))
        sheet[cell_loc] = data
        if num_of_cols > max_merged_cell_len:
            max_merged_cell_len = num_of_cols
        sheet[cell_loc].fill = PatternFill(fill_type='solid', start_color='ffd966')
        sheet[cell_loc].border += Border(left=Side(style='thin'), right=Side(style='thin'))

        # marvel movie checker
        marvel = BeautifulSoup(urlopen("https://pro.imdb.com" + divs[y].find('h4').find('a').get('href')),
                               'html.parser').find(class_="a-size- a-align- a-link-",
                                                   href="https://pro.imdb.com/name/nm0270559/?ref_=tt_pub_fm_prodr")
        if marvel is not None and marvel.text.strip() == "Kevin Feige":
            sheet[cell_loc].fill = PatternFill(fill_type='solid', start_color="ea9999")
        row += 1
    # sets a border at the bottom of each month
    set_cell_border(1, num_of_cols + 1, row - 1, row - 1, 'bottom', True)

    # makes gray cells next to month cell depending on longest merged cell length
    cell_loc = 'B' + str(current_month_row)
    sheet.merge_cells(cell_loc + ':' + get_column_letter(1 + max_merged_cell_len) + str(current_month_row))
    sheet[cell_loc].fill = PatternFill(fill_type='solid', start_color='666666')
    sheet[cell_loc].border += Border(left=Side(style='thin'), right=Side(style='thin'))

    # reiterate through white cells in month to set borders
    for col in sheet.iter_cols(min_col=2, max_col=max_merged_cell_len + 1, min_row=current_month_row + 1, max_row=row - 1):
        for cell in col:
            if not isinstance(cell, MergedCell) and cell.column != 2:
                if isinstance(cell.offset(1, 0), MergedCell): # if merged cell below
                    cell.border += Border(bottom=Side(style='thin'))
                if isinstance(cell.offset(-1, 0), MergedCell): # if merged cell above
                    cell.border += Border(top=Side(style='thin'))

    # check cells above gray cells and if not merged, set border on bottom
    set_cell_border(2, max_merged_cell_len + 1, current_month_row - 1, current_month_row - 1, 'bottom', False)

    print('Found ' + str(row - current_month_row - 1) + ' movies for this month.')
print('Creating a file \'' + file_name + '\' in Python script directory!')
workbook.save(filename=file_name)
